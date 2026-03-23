import hashlib
import os
from typing import Any

import fitz
from openpyxl import load_workbook
from langchain_community.document_loaders import TextLoader
from langchain_core.documents import Document

from backend.utils.logger_handler import logger


def get_file_md5_hex(file_path: str):
    if not os.path.exists(file_path):
        logger.error(f"File {file_path} does not exist")
        return
    if not os.path.isfile(file_path):
        logger.error(f"Path {file_path} is not a file")
        return
    md5_obj = hashlib.md5()

    chunk_size = 4096
    try:
        with open(file_path, "rb") as f:
            while chunk := f.read(chunk_size):
                md5_obj.update(chunk)
            md5_hex = md5_obj.hexdigest()
            return md5_hex
    except Exception as e:
        logger.error(f"[get_file_md5_hex]Error while calculating md5 of file {file_path}: {e}")


def listdir_with_allowed_type(dir_path, allowed_types: tuple[str, ...]):
    files: list[str] = []
    if not os.path.isdir(dir_path):
        logger.error(f"[listdir_with_allowed_type]Path {dir_path} is not a directory")
        return tuple(files)

    for f in os.listdir(dir_path):
        if f.endswith(allowed_types):
            files.append(os.path.join(dir_path, f))

    return tuple(files)


def pdf_loader(file_path: str, password: str | None = None) -> list[Document]:
    """
    使用 PyMuPDF (pymupdf) 抽取文本，对畸形/非标准 xref 的 PDF 通常比 pypdf 更宽容。
    """
    documents: list[Document] = []
    doc: Any = fitz.open(file_path)
    page_count = 0
    try:
        page_count = len(doc)
        if doc.is_encrypted:
            ok = doc.authenticate(password or "")
            if not ok:
                raise ValueError("PDF 已加密，需要提供正确密码")
        for page_index in range(page_count):
            page = doc[page_index]
            try:
                text = (page.get_text("text") or "").strip()
            except Exception as e:
                logger.warning(
                    "PyMuPDF 读取页失败 page=%s file=%s: %s",
                    page_index + 1,
                    file_path,
                    e,
                )
                continue
            if not text:
                continue
            documents.append(
                Document(
                    page_content=text,
                    metadata={
                        "source": file_path,
                        "page": page_index + 1,
                    },
                )
            )
    finally:
        doc.close()
    if not documents and page_count > 0:
        logger.warning(
            "PDF 无可用文本层（可能为扫描件/图片）: %s",
            file_path,
        )
    return documents


def text_loader(file_path: str) -> list[Document]:
    return TextLoader(file_path=file_path, encoding="utf-8").load()


def xlsx_loader(file_path: str) -> list[Document]:
    """
    将 xlsx 每行转为一个 Document；行内单元格用 " | " 拼接。
    全空行跳过；metadata 含 source、sheet、row（工作表内行号，从 1 起）。
    """
    documents: list[Document] = []
    wb = load_workbook(file_path, read_only=True, data_only=True)
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                cells: list[str] = []
                for c in row:
                    if c is None:
                        cells.append("")
                    else:
                        cells.append(str(c).strip())
                if not any(cells):
                    continue
                line = " | ".join(c for c in cells if c)
                if not line:
                    continue
                documents.append(
                    Document(
                        page_content=line,
                        metadata={
                            "source": file_path,
                            "sheet": sheet_name,
                            "row": row_idx,
                        },
                    )
                )
    finally:
        wb.close()
    return documents
